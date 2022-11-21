#!/usr/bin/env python3

import netmiko
from netmiko import ConnectHandler
from getpass import getpass
from openpyxl import load_workbook
import re

username = input('Enter your SSH username: ')
password = getpass()


wb = load_workbook("excel_test.xlsx")
sh1 = wb['Sheet1']

rows = sh1.iter_rows(min_row=1, max_row=12, min_col=1, max_col=5)

ip_addresses = []
ch_number = []

for a,b,c,d,e in rows:
    ip_addresses.append(e.value)

final_list = ip_addresses[1:3]


for indi_device in final_list:
    asa = {
        'device_type': 'cisco_asa',
        'ip': indi_device,
        'username': username,
        'password': password,
    }
    net_connect = ConnectHandler(**asa)
    device_name = net_connect.send_command('show hostname')
    chassis_info = net_connect.send_command('show inventory')

    x = re.search(r"SN:.\w*", chassis_info)
    chassis_serialnum = x.group()
    ch_number.append(chassis_serialnum)

for i in range(2,4):
    for y in ch_number:
        sh1.cell(row=i, column =1).value = y

wb.save("excel_test.xlsx")





